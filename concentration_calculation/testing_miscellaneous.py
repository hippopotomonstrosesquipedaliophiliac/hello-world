import tkinter as tk 
from tkinter import ttk, messagebox, filedialog
from functools import partial
import re
import os
import time
import pythoncom
import win32com.client
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
class miscellaneous():
    def __init__(self):
        self.root = tk.Tk()

    # UNIVERSAL handlers
    def on_ok(self, win_event, result, get_value=None, validate=None,
              error_title="Invalid input", error_msg="Please fix the input."):
        """
        - win_event: the Toplevel to close
        - result: a dict to store output in, e.g. {"value": ...}
        - get_value: optional callable to read current value (e.g. entry.get)
        - validate: optional callable(value) -> bool
        """
        if get_value is not None:
            val = get_value()
            if validate and not validate(val):
                messagebox.showerror(error_title, error_msg)
                return  # keep dialog open
            result["value"] = val
        else:
            # simple confirm-style dialog
            result["value"] = True
        win_event.destroy()

    def on_cancel(self, win_event, result):
        result["value"] = False
        win_event.destroy()


    def to_confirm(self, title="Confirm", label="Are you sure?"):
        win = tk.Toplevel(self.root)
        win.title(title)
        win.resizable(False, False)

        ttk.Label(win, text=label, wraplength=420, justify="left").grid(
            row=0, column=0, columnspan=2, padx=12, pady=12, sticky = "we"
        )
        result = {"value": False}

        ttk.Button(
            win, text="OK",
            command=partial(self.on_ok, win, result)  # no value to read
        ).grid(row=1, column=0, padx=8, pady=(0, 12))

        ttk.Button(
            win, text="Cancel",
            command=partial(self.on_cancel, win, result)
        ).grid(row=1, column=1, padx=8, pady=(0, 12))

        win.bind("<Return>",  lambda e: self.on_ok(win, result))
        win.bind("<Escape>",  lambda e: self.on_cancel(win, result))
        win.protocol("WM_DELETE_WINDOW", partial(self.on_cancel, win, result))

        win.transient(self.root); win.grab_set(); win.update_idletasks()
        x = (win.winfo_screenwidth()-win.winfo_width())//2
        y = (win.winfo_screenheight()-win.winfo_height())//3
        win.geometry(f"+{x}+{y}")
        self.root.wait_window(win)
        return result["value"]

    # Example 2: input dialog using the same universal handlers
    def take_input(self, title="Input", label="Please enter a value:"):
        win = tk.Toplevel(self.root)
        win.title(title); win.resizable(False, False)

        ttk.Label(win, text=label, wraplength=420, justify="left").grid(
            row=0, column=0, padx=12, pady=12, sticky="w"
        )
        value_var = tk.StringVar()
        entry = ttk.Entry(win, width=24, textvariable=value_var)
        entry.grid(row=0, column=1, padx=10, pady=10); entry.focus()
        result = {"value": None}

        get_value = value_var.get
        validate  = lambda s: bool(s.strip())

        ttk.Button(
            win, text="Accept",
            command=partial(self.on_ok, win, result, get_value, validate,
                            "Missing input", "Please enter a non-empty value.")
        ).grid(row=1, column=0, padx=8, pady=(0, 12))

        ttk.Button(
            win, text="Cancel",
            command=partial(self.on_cancel, win, result)
        ).grid(row=1, column=1, padx=8, pady=(0, 12))

        win.bind("<Return>",   lambda e: self.on_ok(win, result, get_value, validate,
                                                    "Missing input", "Please enter a non-empty value."))
        win.bind("<KP_Enter>", lambda e: self.on_ok(win, result, get_value, validate,
                                                    "Missing input", "Please enter a non-empty value."))
        win.bind("<Escape>",   lambda e: self.on_cancel(win, result))
        win.protocol("WM_DELETE_WINDOW", partial(self.on_cancel, win, result))

        win.transient(self.root); win.grab_set(); win.update_idletasks()
        x = (win.winfo_screenwidth()-win.winfo_width())//2
        y = (win.winfo_screenheight()-win.winfo_height())//3
        win.geometry(f"+{x}+{y}")
        self.root.wait_window(win)
        return result["value"]
    def select_from_list(self,options:list = None,title:str ="select an option",label:str = "available options"):
        """
        Show a modal dialog with radio-button options.
        options: list of strings (e.g., file paths or names)
        Returns: the selected string or None.
        """
        win = tk.Toplevel(self.root)
        win.title(title)
        win.resizable(False,False)

        container = ttk.Frame(win)
        container.grid(row=1,column=0,columnspan=2,padx=12,pady=6,sticky="nsew")
        canvas = tk.Canvas(container,width=520, height=240,highlightthickness=0)
        vsb = ttk.Scrollbar(container, orient="vertical",command=canvas.yview)
        inner = ttk.Frame(canvas)
        
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        container.columnconfigure(0, weight=1)
        var = tk.StringVar(value=options[0] if options else "")
        for opt in options:
            ttk.Radiobutton(inner, text=opt, value=opt, variable=var).pack(anchor="w", pady=2)
        result = {"value":None}
        ttk.Button(win, text="OK", command=partial(self.on_ok, win_event=win,result=result,get_value=var.get)).grid(row=2, column=0, padx=12, pady=(6, 12), sticky="e")
        ttk.Button(win, text="Cancel", command=partial(self.on_cancel,win_event=win,result=result,get_value=var.get)).grid(row=2, column=1, padx=12, pady=(6, 12), sticky="w")

        win.bind("<Return>", lambda e: partial(self.on_ok, win_event=win,result=result,get_value=var.get))
        win.bind("<KP_Enter>", lambda e:partial(self.on_ok, win_event=win,result=result,get_value=var.get))
        win.bind("<Escape>", lambda e: partial(self.on_cancel,win_event=win,result=result,get_value=var.get))
        win.protocol("WM_DELETE_WINDOW", partial(self.on_cancel,win_event=win,result=result,get_value=var.get))

        win.transient(self.root); win.grab_set(); win.update_idletasks()
        x = (win.winfo_screenwidth()-win.winfo_width())//2
        y = (win.winfo_screenheight()-win.winfo_height())//3
        win.geometry(f"+{x}+{y}")

        self.root.wait_window(win)
        return result["value"]
    result = {"value": None}
    def close(self):
        self.root.destroy()
    
    @staticmethod
    def clean_equation(equation):
        # Remove standalone numbers (not followed by letters)
        equation = re.sub(r'(?<![A-Za-z])\b\d+\b(?![A-Za-z])', '', equation)
        # Replace multiple pluses with a single plus
        equation = re.sub(r'\++', '+', equation)
        # Remove pluses before/after arrows (=, ⇌, <=>, <->)
        equation = re.sub(r'\+\s*(=|⇌|<=>|<->)', r'\1', equation)
        equation = re.sub(r'(=|⇌|<=>|<->)\s*\+', r'\1', equation)
        # Remove pluses at the start/end
        equation = re.sub(r'^\s*\+\s*', '', equation)
        equation = re.sub(r'\s*\+\s*$', '', equation)
        # Remove empty reactant or product sides (e.g., "+ ⇌" or "+ =")
        equation = re.sub(r'(\s*[\=⇌<=><->]\s*)\+', r'\1', equation)
        equation = re.sub(r'\+(\s*[\=⇌<=><->]\s*)', r'\1', equation)
        # Remove extra spaces
        equation = re.sub(r'\s+', ' ', equation).strip()
        return equation
    @staticmethod
    def choose_output_path(ui, default_name:str="concentration_requesting_data.xlsx", dir:str = None):
        """
        Steps:
        1) Ask to use default name
        - yes  -> name = default_name, go to step 3
        - no   -> go to step 2
        2) Ask user for a file name (loop until valid or cancel)
        3) If file exists:
            - yes -> confirm overwrite
                    - yes: remove old, exit with this name
                    - no : go back to step 2
            - no  -> exit with this name
        Returns: str path or None if user cancels.
        """
        # Step 1
        use_default = ui.to_confirm(
            title="Default file?",
            label=f"Use default name?\n(default: {default_name})"
        )

        if use_default:
            name = os.path.join(dir,default_name)
            # Step 3
            if os.path.exists(name):
                overwrite = ui.to_confirm(
                    title="File already exists",
                    label=f"'{name}' already exists.\nOverwrite?"
                )
                if overwrite:
                    os.remove(name)
                    return name
                else:
                    # Back to step 2
                    pass
            else:
                return name

        # Step 2 (loop until we get a usable name or user cancels)
        while True:
            val = ui.take_input(
                title="Output filename",
                label="Enter a file name for the reactants' concentration dataframe:"
            )
            if val is None:          # user canceled
                return None

            name = os.path.join(dir,val.strip())
            if not name:
                ui.messagebox.showerror("Invalid input", "File name cannot be empty.")
                continue
            if not name.lower().endswith(".xlsx"):
                name += ".xlsx"

            # Step 3
            if os.path.exists(name):
                overwrite = ui.to_confirm(
                    title="File already exists",
                    label=f"'{name}' already exists.\nOverwrite?"
                )
                if overwrite:
                    os.remove(name)
                    return name       # exit block, keep going
                else:
                    continue          # back to step 2
            else:
                return name           # file doesn't exist -> done



def check_if_previous_data_frame_exist(folders_path : str = None):
    folders = Path(folders_path)
    list_of_previous_name_from_file = []
    for item in folders.glob("*.xlsx"):
        species_concentration_name_list = []
        if item.name.startswith("~$"):
            continue
        data_sheets_name = load_workbook(item, read_only=True).sheetnames  
        for individual_sheet in data_sheets_name:
            df_filled = pd.read_excel(item, sheet_name=f"{individual_sheet}")
            species_concentration_name_list.extend([prev_spec_name for prev_spec_name in df_filled["species"]])        
        list_of_previous_name_from_file.append({os.path.basename(item): sorted(species_concentration_name_list)})
    return list_of_previous_name_from_file
def edit_then_wait(path, visible=True, require_saved=True, timeout=None, poll=0.5):
    """Open `path` in Excel (COM), show it, and block until user closes it.
       If `require_saved` is True, ensure it’s saved before we accept closure."""
    pythoncom.CoInitialize()
    excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = bool(visible)

    # Open the workbook we will track
    wb = excel.Workbooks.Open(os.path.abspath(path))

    # Let the user edit in Excel’s UI
    start = time.time()
    while True:
        # Try to see if wb is still open by checking if it's still in the collection
        still_open = False
        for w in list(excel.Workbooks):
            try:
                if os.path.abspath(w.FullName).lower() == os.path.abspath(path).lower():
                    still_open = True
                    # optional save requirement
                    if require_saved and not bool(w.Saved):
                        break  # still open & unsaved → keep waiting
                    break
            except Exception:
                pass

        if not still_open:
            # The user closed it. Done.
            break

        if timeout is not None and (time.time() - start) > timeout:
            return False  # timed out
        time.sleep(poll)

    # Do NOT quit Excel here; the user may have other workbooks open.
    return True



